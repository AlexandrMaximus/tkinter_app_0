from openpyxl import load_workbook
from tkinter import *
from tkinter import messagebox
from tkinter import ttk



#fn = 'example.xlsx'
#wb = load_workbook(fn)
#ws = wb['data']

#for row in range(1, 4):
    #for col in range(1, 5):
        #value =str(row) + str(col)
        #cell = ws.cell(row=row, column=col)
        #cell.value = value

#wb.save(fn)
#wb.close()

def selected(event):
    # получаем выделенный элемент
    selection = e1.get()
    print(selection)
    lb1["text"] = f"вы выбрали: {selection}"

def save():
    fn = 'example.xlsx'
    fn2 = 'example2.xlsx'
    wb = load_workbook(fn)
    ws = wb['data']
    wb2 = load_workbook(fn2)
    ws2 = wb2['data']
    data = (e1.get(), e2.get(), e3.get(), e4.get(), e5.get(), e6.get(), e7.get())
    ws.append(data)
    wb.save(fn)
    wb.close()
    ws2.append(data)
    wb2.save(fn2)
    wb2.close()
    messagebox.askokcancel('сохранение', 'успешно сохранено')

root = Tk()
root.title('Тест')
root.geometry('200x200')

plases = ["M460", "LV800", "HD2200", "L250"]
e1 = ttk.Combobox(values=plases, state="readonly")
e1.pack()
e1.bind("<<ComboboxSelected>>", selected)
#e1 = Entry(root)
#e1.pack()
lb1 = Label(root, text='участок',font='Arial 15 bold')
lb1.pack()

e2 = Entry(root)
e2.pack()
lb2 = Label(root, text='фамилия',font='Arial 15 bold')
lb2.pack()

e3 = Entry(root)
e3.pack()
lb3 = Label(root, text='шифр детали',font='Arial 15 bold')
lb3.pack()

e4 = Entry(root)
e4.pack()
lb4 = Label(root, text='наименование детали',font='Arial 15 bold')
lb4.pack()

e5 = Entry(root)
e5.pack()
lb5 = Label(root, text='операция',font='Arial 15 bold')
lb5.pack()

e6 = Entry(root)
e6.pack()
lb6 = Label(root, text='количество',font='Arial 15 bold')
lb6.pack()

e7 = Entry(root)
e7.pack()
lb7 = Label(root, text='норма',font='Arial 15 bold')
lb7.pack()

btn = Button(root, text='Save', font='Arial 15 bold', command= save)
btn.pack()


root.mainloop()